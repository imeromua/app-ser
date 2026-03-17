"""
exporter.py — експорт у Excel для детального, сумарного та документального звітів.
"""

import io

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


def export_excel(header, rows, grand, report_type='detail'):
    """
    Експортує дані у Excel.
    report_type='detail'   — детальний звіт (з колонкою Місяць)
    report_type='summary'  — сумарний звіт (без колонки Місяць)
    report_type='document' — звіт по документах (Дата, Тип операції, ...)
    """
    out = io.BytesIO()

    if report_type == 'summary':
        cols     = ['Артикул', 'Назва', 'ПрВ (прихід)', 'Кнк (продажі)',
                    'ПрИ (переміщення)', 'СпП (списання)', 'Апс (акт пересорту)',
                    'Залишок', 'Ціна', 'Сума']
        col_keys = ['Артикул', 'Назва', 'ПрВ', 'Кнк', 'ПрИ', 'СпП', 'Апс',
                    'Залишок', 'Ціна', 'Сума']
        col_widths = [12, 42, 10, 10, 14, 13, 16, 10, 10, 13]
    elif report_type == 'document':
        cols     = ['Артикул', 'Назва', 'Дата', 'Тип операції', 'Документ',
                    'Прихід', 'Розхід', 'Кількість', 'Залишок']
        col_keys = ['Артикул', 'Назва', 'Дата', 'Операція', 'Документ',
                    'Прихід', 'Розхід', 'Кількість', 'Залишок']
        col_widths = [12, 42, 12, 22, 48, 10, 10, 10, 10]
    else:
        cols     = ['Артикул', 'Назва', 'Місяць', 'ПрВ (прихід)', 'Кнк (продажі)',
                    'ПрИ (переміщення)', 'СпП (списання)', 'Апс (акт пересорту)',
                    'Залишок', 'Ціна', 'Сума']
        col_keys = ['Артикул', 'Назва', 'Місяць', 'ПрВ', 'Кнк', 'ПрИ', 'СпП', 'Апс',
                    'Залишок', 'Ціна', 'Сума']
        col_widths = [12, 42, 10, 10, 10, 14, 13, 16, 10, 10, 13]

    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    sfill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    sfont = Font(bold=True, color='1F3864')

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        pd.DataFrame([[]]).to_excel(writer, index=False, header=False,
                                    sheet_name='Рух товарів', startrow=0)
        wb = writer.book
        ws = writer.sheets['Рух товарів']

        ws['A1'].value = header.get('title', '')
        ws['A1'].font  = Font(bold=True, size=14, color='C0392B')
        ws['A2'].value = 'Магазин:';  ws['D2'].value = header.get('shop', '')
        ws['K2'].value = header.get('period', '')
        ws['A4'].value = 'Склад:';    ws['D4'].value = header.get('warehouse', '')

        HR = 6
        for ci, cn in enumerate(cols, 1):
            c = ws.cell(row=HR, column=ci, value=cn)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal='center')

        dr = HR + 1
        for row in rows:
            rt = row.get('type')
            if rt == 'spacer':
                dr += 1
                continue
            for ci, ck in enumerate(col_keys, 1):
                val  = row.get(ck, '')
                cell = ws.cell(row=dr, column=ci)
                if val == '' or val is None:
                    cell.value = None
                elif isinstance(val, str):
                    cell.value = val
                else:
                    try:
                        float_cols = {'Ціна', 'Сума', 'Прихід', 'Розхід', 'Кількість', 'Залишок'}
                        cell.value = float(val) if ck in float_cols else int(val)
                        if ck == 'Сума':
                            cell.number_format = '#,##0.00'
                        if ck == 'Ціна':
                            cell.number_format = '0.00'
                    except Exception:
                        cell.value = val
                if rt in ('subtotal', 'summary'):
                    cell.fill = sfill
                    cell.font = sfont
                if ci > 2:
                    cell.alignment = Alignment(horizontal='right')
            dr += 1

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f'A{HR + 1}'

        # ── Grand Total row ───────────────────────────────────────────────
        gfill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        gfont = Font(bold=True, color='FFFFFF', size=12)

        if report_type == 'summary':
            grand_vals = ['', 'ВСЬОГО', grand['ПрВ'], grand['Кнк'],
                          grand['ПрИ'], grand['СпП'], grand['Апс'],
                          grand['Залишок'], '', grand['Сума']]
        elif report_type == 'document':
            grand_vals = ['', 'ВСЬОГО', '', '', '',
                          grand.get('Прихід', ''), grand.get('Розхід', ''),
                          '', grand.get('Залишок', '')]
        else:
            grand_vals = ['', 'ВСЬОГО', '', grand['ПрВ'], grand['Кнк'],
                          grand['ПрИ'], grand['СпП'], grand['Апс'],
                          grand['Залишок'], '', grand['Сума']]

        for ci, val in enumerate(grand_vals, 1):
            c = ws.cell(row=dr, column=ci, value=val if val != '' else None)
            c.fill = gfill
            c.font = gfont
            if ci > 2:
                c.alignment = Alignment(horizontal='right')
            if ci == len(grand_vals) and report_type != 'document':
                c.number_format = '#,##0.00'

    out.seek(0)
    return out
