#!/usr/bin/env python3
"""
Рух товарів — веб-аналізатор
Запуск: python app.py  →  http://localhost:5000
"""

from dotenv import load_dotenv
load_dotenv()

import hmac
import io
import logging
import os
import secrets
import threading
import uuid

import openpyxl
from openpyxl.styles import Font as XlFont, Alignment as XlAlign, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from celery.result import AsyncResult
from functools import wraps

from flask import Flask, request, render_template, send_file, session, jsonify, Response as FlaskResponse, redirect, url_for
from werkzeug.utils import secure_filename
import pandas as pd

from categories import detect_category, CATEGORIES
from session_store import save_session_data, load_session_data, cleanup_old_sessions
from parser import parse_xls, op_display_name
from builder import build_rows, build_summary_rows, build_document_rows
from exporter import export_excel
from tasks import celery, generate_pdf_task  # noqa: F401 — celery app must be imported
from importer import run_import
from db import get_conn

logging.basicConfig(level=logging.WARNING)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        expected_login = os.environ.get('APP_LOGIN', '')
        expected_password = os.environ.get('APP_PASSWORD', '')
        provided_login = request.form.get('login', '')
        provided_password = request.form.get('password', '')
        login_ok = hmac.compare_digest(provided_login, expected_login)
        password_ok = hmac.compare_digest(provided_password, expected_password)
        if login_ok and password_ok:
            session['logged_in'] = True
            next_url = request.args.get('next', '')
            if next_url and next_url.startswith('/') and not next_url.startswith('//'):
                return redirect(next_url)
            return redirect(url_for('index'))
        error = 'Невірний логін або пароль'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/robots.txt')
def robots_txt():
    return app.send_static_file('robots.txt')


@app.route('/')
@login_required
def index():
    return render_template('index.html', error=None)


@app.route('/upload', methods=['POST'])
@login_required
def upload():
    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        return render_template('index.html', error='Файл не знайдено')
    try:
        report_type = request.form.get('report_type', 'detail')
        all_ops, all_prices = [], {}
        first_header = None

        for f in files:
            if not f.filename:
                continue
            buf = io.BytesIO(f.read())
            result = parse_xls(buf)
            hdr = result['header']
            if first_header is None:
                first_header = hdr

            all_prices.update({
                a['article_id']: a['price']
                for a in result['articles']
                if a.get('price')
            })

            articles_map = {a['article_id']: a for a in result['articles']}
            records = []
            for op in result['operations']:
                art = articles_map.get(op['article_id'], {})
                d   = op['op_date']
                ym  = f"{d.year}-{d.month:02d}" if d else ''
                qty = op['qty']
                src = op['col_source']
                records.append({
                    'Артикул':    op['article_id'],
                    'Назва':      art.get('name', ''),
                    'Операція':   op_display_name(
                                      op['doc_type'],
                                      op.get('subdoc_type'),
                                      op.get('direction'),
                                  ),
                    'Документ':   op.get('doc_type', '') + '/' + op.get('doc_code', ''),
                    'Піддокумент': (
                        op.get('subdoc_type', '') + '/' + op.get('subdoc_code', '')
                        if op.get('subdoc_type') and op.get('subdoc_code')
                        else ''
                    ),
                    'Дата':       d,
                    'Рік-Місяць': ym,
                    'Кількість':  qty,
                    'Прихід':     qty if src == 'G' and qty > 0 else (qty if src == 'I' and qty > 0 else 0.0),
                    'Розхід':     abs(qty) if src == 'H' else (abs(qty) if src == 'I' and qty < 0 else 0.0),
                })
            ops_df = pd.DataFrame(records) if records else pd.DataFrame()
            all_ops.append(ops_df)

        combined_df = pd.concat(all_ops, ignore_index=True, sort=False) if all_ops else pd.DataFrame()
        header = first_header or {}

        if report_type == 'summary':
            rows, grand = build_summary_rows(combined_df, all_prices)
        elif report_type == 'document':
            rows, grand = build_document_rows(combined_df, all_prices)
        else:
            rows, grand = build_rows(combined_df, all_prices)

        if report_type == 'document':
            all_names = [r['Назва'] for r in rows if r.get('type') == 'doc_data']
            art_count = len(set(r['Артикул'] for r in rows if r.get('type') == 'doc_data'))
            row_count = sum(1 for r in rows if r.get('type') == 'doc_data')
        else:
            all_names = [r['Назва'] for r in rows if r.get('type') in ('data', 'summary')]
            art_count = len(set(
                r['Артикул'] for r in rows
                if r.get('type') in ('subtotal', 'summary')
            ))
            row_count = sum(1 for r in rows if r.get('type') in ('data', 'summary'))

        category = detect_category(all_names)
        safe_category = category.replace('/', '-').replace(' ', '_')

        if report_type == 'summary':
            download_name = f"{safe_category}_сумарний_звіт.xlsx"
        elif report_type == 'document':
            download_name = f"{safe_category}_по_документах.xlsx"
        else:
            download_name = f"{safe_category}_детальний_звіт.xlsx"

        session_id = save_session_data({
            'header': header,
            'rows': rows,
            'grand': grand,
            'filename': download_name,
            'category': category,
            'report_type': report_type,
        })
        session['sid'] = session_id
        threading.Thread(target=cleanup_old_sessions, daemon=True).start()

        return render_template('result.html',
                               header=header, rows=rows, grand=grand,
                               art_count=art_count, row_count=row_count,
                               category=category, report_type=report_type)
    except Exception as e:
        logging.exception('Error processing uploaded file')
        return render_template('index.html', error=f'Помилка обробки файлу: {e}')


@app.route('/download')
@login_required
def download():
    sid = session.get('sid')
    data = load_session_data(sid)
    if not data:
        return 'Немає даних або сесія застаріла', 400
    header        = data['header']
    rows          = data['rows']
    grand         = data['grand']
    download_name = data.get('filename', 'ruh_tovariv_звіт.xlsx')
    report_type   = data.get('report_type', 'detail')
    buf = export_excel(header, rows, grand, report_type=report_type)
    return send_file(buf, as_attachment=True,
                     download_name=download_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download_inventory')
@login_required
def download_inventory():
    """
    Відомість інвентаризації з артикулів ПОТОЧНОГО звіту (з сесії).
    Не запитує БД — використовує тільки дані сесії поточного звіту.
    """
    sid = session.get('sid')
    data = load_session_data(sid)
    if not data:
        return 'Немає даних або сесія застаріла', 400

    rows = data.get('rows', [])
    report_type = data.get('report_type', 'detail')
    category = data.get('category', 'товари')
    hdr = data.get('header', {})
    safe_category = category.replace('/', '-').replace(' ', '_')

    # Беремо артикули виключно з поточного звіту (сесія)
    if report_type == 'document':
        article_map: dict = {}
        for r in rows:
            if r.get('type') == 'doc_data':
                article_map[r['Артикул']] = r
        inv_rows = list(article_map.values())
    else:
        inv_rows = [r for r in rows if r.get('type') in ('subtotal', 'summary')]

    if not inv_rows:
        return 'Немає даних для відомості інвентаризації', 400

    inv_rows.sort(key=lambda r: r.get('Назва', '').lower())

    # ── Build xlsx ──────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Інвентаризація'

    NUM_COLS = 8
    col_widths = [5, 12, 56, 10, 14, 16, 30, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    title_font   = XlFont(bold=True, color='FF0000', size=14)
    label_font   = XlFont(bold=True)
    value_font   = XlFont(color='1F3864')
    hdr_fill     = PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid')
    hdr_font     = XlFont(bold=True)
    hdr_align    = XlAlign(horizontal='center', vertical='center', wrap_text=True)
    thin         = Side(style='thin')
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bot_border   = Border(bottom=thin)
    center_align = XlAlign(horizontal='center', vertical='center')
    right_align  = XlAlign(horizontal='right', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    tc = ws.cell(row=1, column=1, value='Відомість інвентаризації')
    tc.font = title_font
    tc.alignment = XlAlign(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    c = ws.cell(row=2, column=1, value='Маркет:')
    c.font = label_font
    c.alignment = XlAlign(vertical='center')
    c = ws.cell(row=2, column=3, value=hdr.get('shop', ''))
    c.font = value_font
    c.alignment = XlAlign(vertical='center', wrap_text=True)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=NUM_COLS)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    c = ws.cell(row=3, column=1, value='Єрархія:')
    c.font = label_font
    c.alignment = XlAlign(vertical='center')
    c = ws.cell(row=3, column=3, value=hdr.get('warehouse', ''))
    c.font = value_font
    c.alignment = XlAlign(vertical='center', wrap_text=True)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=NUM_COLS)

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    c = ws.cell(row=4, column=1, value='Примітки:')
    c.font = label_font
    c.alignment = XlAlign(vertical='center')
    c = ws.cell(row=4, column=3, value=f'Позапланова інв_{category}')
    c.font = value_font
    c.alignment = XlAlign(vertical='center', wrap_text=True)
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=NUM_COLS)

    ws.row_dimensions[5].height = 8

    HEADER_ROW = 6
    col_headers = [
        '№', 'Артикул', 'Назва',
        'Од.\nвим.', 'База\n(залишок)',
        'Фактичні\nзалишки', 'Примітки', 'Час\nінвентаризації',
    ]
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = cell_border
    ws.row_dimensions[HEADER_ROW].height = 42

    DATA_START = HEADER_ROW + 1
    for i, r in enumerate(inv_rows, 1):
        dr = DATA_START + i - 1

        c = ws.cell(row=dr, column=1, value=i)
        c.border = cell_border
        c.alignment = center_align

        c = ws.cell(row=dr, column=2, value=r.get('Артикул', ''))
        c.border = cell_border
        c.alignment = center_align
        c.font = XlFont(bold=True)

        c = ws.cell(row=dr, column=3, value=r.get('Назва', ''))
        c.border = cell_border
        c.alignment = XlAlign(vertical='center', wrap_text=True)

        ws.cell(row=dr, column=4).border = cell_border

        zal = r.get('Залишок', '')
        c = ws.cell(row=dr, column=5)
        c.border = cell_border
        c.alignment = right_align
        c.font = XlFont(bold=True)
        if zal != '' and zal is not None:
            try:
                fval = float(zal)
                if fval.is_integer():
                    c.value = int(fval)
                else:
                    c.value = round(fval, 2)
                    c.number_format = '0.##'
            except (ValueError, TypeError):
                c.value = zal

        ws.cell(row=dr, column=6).border = cell_border
        ws.cell(row=dr, column=7).border = cell_border
        ws.cell(row=dr, column=8).border = cell_border

    last_data_row = DATA_START + len(inv_rows) - 1
    fr = last_data_row + 2

    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=NUM_COLS)
    ws.cell(row=fr, column=1, value='Особи, які проводили перерахунок:')
    ws.row_dimensions[fr].height = 22

    for sig_idx in range(2):
        line_row  = fr + 2 + sig_idx * 4
        label_row = line_row + 1
        for col_idx in range(3, 6):
            ws.cell(row=line_row, column=col_idx).border = bot_border
        ws.merge_cells(start_row=label_row, start_column=3, end_row=label_row, end_column=5)
        pip = ws.cell(row=label_row, column=3, value='(ПІП)')
        pip.alignment = center_align
        for col_idx in range(7, 9):
            ws.cell(row=line_row, column=col_idx).border = bot_border
        ws.merge_cells(start_row=label_row, start_column=7, end_row=label_row, end_column=8)
        sign = ws.cell(row=label_row, column=7, value='(підпис)')
        sign.alignment = center_align

    nachal_row = fr + 2 + 2 * 4 + 1
    ws.merge_cells(start_row=nachal_row, start_column=1, end_row=nachal_row, end_column=NUM_COLS)
    ws.cell(row=nachal_row, column=1,
            value='Начальник відділу  ___________________________________')
    ws.row_dimensions[nachal_row].height = 22

    date_row = nachal_row + 2
    ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=NUM_COLS)
    ws.cell(row=date_row, column=1,
            value='Дата проведення  ________________________  час з  ________________  по  ________________')
    ws.row_dimensions[date_row].height = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f'{safe_category}_інвентаризація.xlsx'
    return send_file(out, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download_pdf')
@login_required
def download_pdf():
    return (
        'Цей маршрут більше не підтримується. '
        'Використовуйте /export/pdf/start → /export/pdf/status → /export/pdf/result.',
        410,
    )


# ── Async PDF export via Celery ──────────────────────────────────────────────

@app.route('/export/pdf/start', methods=['POST'])
@login_required
def export_pdf_start():
    """Запускає фонову задачу генерації PDF. Повертає {"task_id": "..."}."""
    sid = session.get('sid')
    data = load_session_data(sid)
    if not data:
        return jsonify({'error': 'Немає даних або сесія застаріла'}), 400

    html_string = render_template(
        'report_pdf.html',
        header=data['header'],
        rows=data['rows'],
        grand=data['grand'],
        report_type=data.get('report_type', 'detail'),
        category=data.get('category', ''),
    )
    pdf_name = data.get('filename', 'звіт').replace('.xlsx', '.pdf')

    task = generate_pdf_task.delay(html_string, pdf_name)
    return jsonify({'task_id': task.id})


@app.route('/export/pdf/status/<task_id>')
@login_required
def export_pdf_status(task_id):
    """Повертає стан задачі: pending / ready / error."""
    try:
        uuid.UUID(task_id)
    except ValueError:
        return jsonify({'status': 'error', 'error': 'Невалідний task_id'}), 400

    result = AsyncResult(task_id, app=celery)
    state = result.state

    if state in ('PENDING', 'STARTED'):
        return jsonify({'status': 'pending'})
    if state == 'SUCCESS':
        return jsonify({'status': 'ready'})
    error_msg = str(result.result) if result.result else state
    return jsonify({'status': 'error', 'error': error_msg})


@app.route('/export/pdf/result/<task_id>')
@login_required
def export_pdf_result(task_id):
    """Повертає готовий PDF-файл, якщо задача завершена."""
    try:
        uuid.UUID(task_id)
    except ValueError:
        return jsonify({'error': 'Невалідний task_id'}), 400

    result = AsyncResult(task_id, app=celery)
    if result.state != 'SUCCESS':
        return jsonify({'status': 'pending'}), 202

    task_result = result.result
    if not task_result or not isinstance(task_result, dict):
        return jsonify({'error': 'Некоректний результат задачі'}), 500

    pdf_path = task_result.get('path', '')
    pdf_name = task_result.get('filename', 'звіт.pdf')

    if not pdf_path or not os.path.isfile(pdf_path):
        return jsonify({'error': 'Файл не знайдено'}), 404

    from urllib.parse import quote
    encoded_name = quote(pdf_name, safe='')
    content_disposition = f"attachment; filename*=UTF-8''{encoded_name}"

    def stream_and_delete():
        try:
            with open(pdf_path, 'rb') as f:
                yield from iter(lambda: f.read(65536), b'')
        finally:
            try:
                os.unlink(pdf_path)
            except OSError:
                pass

    response = FlaskResponse(
        stream_and_delete(),
        mimetype='application/pdf',
        headers={'Content-Disposition': content_disposition},
    )
    return response


# ── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


# ── Import to DB ─────────────────────────────────────────────────────────────

@app.route('/import', methods=['POST'])
@login_required
def import_to_db():
    """Імпортує XLS-файл до PostgreSQL. Повертає JSON зі статистикою."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'Файл не знайдено'}), 400
    try:
        buf = io.BytesIO(f.read())
        filename = secure_filename(f.filename)
        result = run_import(buf, filename)

        # Серіалізуємо invalid_snapshots (може містити Decimal/date)
        invalid = []
        for row in result.get('invalid_snapshots', []):
            invalid.append({k: str(v) for k, v in row.items()})

        header_info = {}
        try:
            buf.seek(0)
            parsed = parse_xls(buf)
            hdr = parsed.get('header', {})
            header_info = {
                'period_from': str(hdr['period_from']) if hdr.get('period_from') else None,
                'period_to':   str(hdr['period_to'])   if hdr.get('period_to')   else None,
            }
        except Exception:
            pass

        return jsonify({
            'strategy':          result.get('strategy'),
            'ops_inserted':      result.get('ops_inserted', 0),
            'articles_count':    result.get('articles_count', 0),
            'invalid_snapshots': invalid,
            'invalid_count':     len(invalid),
            **header_info,
        })
    except Exception as e:
        logging.exception('Error during DB import')
        return jsonify({'error': f'Помилка імпорту: {e}'}), 500


@app.route('/imports')
@login_required
def list_imports():
    """Повертає список останніх 20 імпортів з таблиці uploads (JSON)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT upload_id, filename, shop, warehouse,
                           period_from, period_to, uploaded_at,
                           strategy, ops_inserted
                    FROM uploads
                    ORDER BY uploaded_at DESC
                    LIMIT 20
                    """
                )
                rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'upload_id':    str(r['upload_id']),
                'filename':     r['filename'],
                'shop':         r['shop'],
                'warehouse':    r['warehouse'],
                'period_from':  str(r['period_from'])  if r['period_from']  else None,
                'period_to':    str(r['period_to'])    if r['period_to']    else None,
                'uploaded_at':  str(r['uploaded_at'])  if r['uploaded_at']  else None,
                'strategy':     r['strategy'],
                'ops_inserted': r['ops_inserted'],
            })
        return jsonify(result)
    except Exception as e:
        logging.exception('Error fetching imports list')
        return jsonify({'error': str(e)}), 500


# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/api/db_status')
@login_required
def api_db_status():
    """Статистика БД: кількість артикулів, операцій, суми по типах операцій."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT COUNT(*) AS cnt FROM articles')
                articles_count = cur.fetchone()['cnt']

                cur.execute('SELECT COUNT(*) AS cnt FROM operations')
                ops_count = cur.fetchone()['cnt']

                cur.execute('SELECT MIN(op_date) AS mn, MAX(op_date) AS mx FROM operations')
                dates = cur.fetchone()
                date_min = str(dates['mn']) if dates['mn'] else None
                date_max = str(dates['mx']) if dates['mx'] else None

                days_in_db = 0
                if dates['mn'] and dates['mx']:
                    days_in_db = (dates['mx'] - dates['mn']).days

                cur.execute(
                    """
                    SELECT doc_type, SUM(ABS(qty)) AS total_qty
                    FROM operations
                    GROUP BY doc_type
                    ORDER BY doc_type
                    """
                )
                by_type = {r['doc_type']: float(r['total_qty']) for r in cur.fetchall()}

                cur.execute(
                    """
                    SELECT SUM(o.qty * a.price) AS total_sum
                    FROM operations o
                    JOIN articles a USING (article_id)
                    WHERE a.price IS NOT NULL
                    """
                )
                row = cur.fetchone()
                total_sum = float(row['total_sum']) if row and row['total_sum'] else 0.0

                cur.execute(
                    """
                    SELECT uploaded_at FROM uploads
                    ORDER BY uploaded_at DESC LIMIT 1
                    """
                )
                last_import_row = cur.fetchone()
                last_import = str(last_import_row['uploaded_at']) if last_import_row else None

                cur.execute(
                    """
                    SELECT a.article_id, a.name, SUM(o.qty) AS balance,
                           a.price, SUM(o.qty) * a.price AS balance_sum
                    FROM operations o
                    JOIN articles a USING (article_id)
                    WHERE a.price IS NOT NULL
                    GROUP BY a.article_id, a.name, a.price
                    HAVING SUM(o.qty) > 0
                    ORDER BY SUM(o.qty) * a.price DESC NULLS LAST
                    LIMIT 10
                    """
                )
                top_articles = []
                for r in cur.fetchall():
                    top_articles.append({
                        'article_id':   r['article_id'],
                        'name':         r['name'],
                        'balance':      float(r['balance']) if r['balance'] else 0.0,
                        'price':        float(r['price'])   if r['price']   else 0.0,
                        'balance_sum':  float(r['balance_sum']) if r['balance_sum'] else 0.0,
                    })

        return jsonify({
            'articles_count': articles_count,
            'ops_count':      ops_count,
            'days_in_db':     days_in_db,
            'total_sum':      total_sum,
            'date_min':       date_min,
            'date_max':       date_max,
            'last_import':    last_import,
            'by_type':        by_type,
            'top_articles':   top_articles,
        })
    except Exception as e:
        logging.exception('Error fetching DB status')
        return jsonify({'error': str(e)}), 500


@app.route('/api/categories')
@login_required
def api_categories():
    """Повертає список категорій з кількістю артикулів з БД."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT article_id, name FROM articles ORDER BY name')
                articles = cur.fetchall()

        category_map: dict = {cat: [] for cat in CATEGORIES}
        category_map['Змішана продукція'] = []

        for art in articles:
            cat = detect_category([art['name']])
            if cat not in category_map:
                category_map[cat] = []
            category_map[cat].append(art['article_id'])

        result = []
        for cat, ids in sorted(category_map.items()):
            if ids:
                result.append({'category': cat, 'count': len(ids)})

        return jsonify(result)
    except Exception as e:
        logging.exception('Error fetching categories')
        return jsonify({'error': str(e)}), 500


@app.route('/api/uploads')
@login_required
def api_uploads():
    """Повертає останні 20 імпортів (JSON для дашборду)."""
    return list_imports()


# ── Reports from DB ───────────────────────────────────────────────────────────

@app.route('/reports_db', methods=['GET'])
@login_required
def reports_db():
    return render_template('dashboard.html', active_tab='reports')


@app.route('/export_db', methods=['POST'])
@login_required
def export_db():
    """Звіт з БД → XLSX файл."""
    from reports import get_summary_report
    if request.is_json:
        body = request.get_json(silent=True) or {}
        date_from = body.get('date_from') or request.form.get('date_from')
        date_to   = body.get('date_to')   or request.form.get('date_to')
    else:
        date_from = request.form.get('date_from')
        date_to   = request.form.get('date_to')

    if not date_from or not date_to:
        return jsonify({'error': 'Потрібно вказати date_from та date_to'}), 400

    try:
        rows = get_summary_report(date_from, date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Звіт'
        headers = ['Артикул', 'Назва', 'Ціна', 'Прихід', 'Продажі',
                   'Списання', 'Переміщення', 'Інвентаризація', 'Залишок', 'Сума залишку']
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h).font = XlFont(bold=True)
        for ri, r in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=r['article_id'])
            ws.cell(row=ri, column=2, value=r['name'])
            ws.cell(row=ri, column=3, value=float(r['price']) if r.get('price') else '')
            ws.cell(row=ri, column=4, value=float(r['total_in']) if r.get('total_in') else 0)
            ws.cell(row=ri, column=5, value=float(r['total_sales']) if r.get('total_sales') else 0)
            ws.cell(row=ri, column=6, value=float(r['total_writeoff']) if r.get('total_writeoff') else 0)
            ws.cell(row=ri, column=7, value=float(r['total_transfer']) if r.get('total_transfer') else 0)
            ws.cell(row=ri, column=8, value=float(r['total_inv']) if r.get('total_inv') else 0)
            ws.cell(row=ri, column=9, value=float(r['balance']) if r.get('balance') else 0)
            ws.cell(row=ri, column=10, value=float(r['balance_sum']) if r.get('balance_sum') else 0)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'звіт_{date_from}_{date_to}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logging.exception('Error exporting DB report')
        return jsonify({'error': str(e)}), 500


# ── Inventory from DB ─────────────────────────────────────────────────────────

@app.route('/download_inventory_db')
@login_required
def download_inventory_db():
    """Відомість інвентаризації з БД по категорії."""
    from reports import get_inventory_template
    category = request.args.get('category', '')

    try:
        all_rows = get_inventory_template()
    except Exception as e:
        logging.exception('Error fetching inventory from DB')
        return f'Помилка отримання даних з БД: {e}', 500

    if not all_rows:
        return 'База даних порожня', 400

    if category:
        inv_rows = [
            r for r in all_rows
            if detect_category([r['Назва']]) == category
        ]
    else:
        inv_rows = all_rows

    if not inv_rows:
        return f'Немає даних для категорії «{category}»', 400

    inv_rows.sort(key=lambda r: r.get('Назва', '').lower())
    safe_category = (category or 'всі').replace('/', '-').replace(' ', '_')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Інвентаризація'
    NUM_COLS = 8
    col_widths = [5, 12, 56, 10, 14, 16, 30, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    title_font   = XlFont(bold=True, color='FF0000', size=14)
    label_font   = XlFont(bold=True)
    value_font   = XlFont(color='1F3864')
    hdr_fill     = PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid')
    hdr_font     = XlFont(bold=True)
    hdr_align    = XlAlign(horizontal='center', vertical='center', wrap_text=True)
    thin         = Side(style='thin')
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bot_border   = Border(bottom=thin)
    center_align = XlAlign(horizontal='center', vertical='center')
    right_align  = XlAlign(horizontal='right', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    tc = ws.cell(row=1, column=1, value='Відомість інвентаризації')
    tc.font = title_font
    tc.alignment = XlAlign(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    c = ws.cell(row=2, column=1, value='Категорія:')
    c.font = label_font
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=NUM_COLS)
    c = ws.cell(row=2, column=3, value=category or 'Всі категорії')
    c.font = value_font

    ws.row_dimensions[3].height = 8

    HEADER_ROW = 4
    col_headers = ['№', 'Артикул', 'Назва', 'Од.\nвим.', 'База\n(залишок)',
                   'Фактичні\nзалишки', 'Примітки', 'Час\nінвентаризації']
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = cell_border
    ws.row_dimensions[HEADER_ROW].height = 42

    DATA_START = HEADER_ROW + 1
    for i, r in enumerate(inv_rows, 1):
        dr = DATA_START + i - 1
        c = ws.cell(row=dr, column=1, value=i)
        c.border = cell_border
        c.alignment = center_align
        c = ws.cell(row=dr, column=2, value=r.get('Артикул', ''))
        c.border = cell_border
        c.alignment = center_align
        c.font = XlFont(bold=True)
        c = ws.cell(row=dr, column=3, value=r.get('Назва', ''))
        c.border = cell_border
        c.alignment = XlAlign(vertical='center', wrap_text=True)
        ws.cell(row=dr, column=4).border = cell_border
        zal = r.get('Залишок', '')
        c = ws.cell(row=dr, column=5)
        c.border = cell_border
        c.alignment = right_align
        c.font = XlFont(bold=True)
        if zal not in ('', None):
            try:
                fval = float(zal)
                c.value = int(fval) if fval.is_integer() else round(fval, 2)
            except (ValueError, TypeError):
                c.value = zal
        ws.cell(row=dr, column=6).border = cell_border
        ws.cell(row=dr, column=7).border = cell_border
        ws.cell(row=dr, column=8).border = cell_border

    last_data_row = DATA_START + len(inv_rows) - 1
    fr = last_data_row + 2
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=NUM_COLS)
    ws.cell(row=fr, column=1, value='Особи, які проводили перерахунок:')
    ws.row_dimensions[fr].height = 22

    for sig_idx in range(2):
        line_row  = fr + 2 + sig_idx * 4
        label_row = line_row + 1
        for col_idx in range(3, 6):
            ws.cell(row=line_row, column=col_idx).border = bot_border
        ws.merge_cells(start_row=label_row, start_column=3, end_row=label_row, end_column=5)
        pip = ws.cell(row=label_row, column=3, value='(ПІП)')
        pip.alignment = center_align
        for col_idx in range(7, 9):
            ws.cell(row=line_row, column=col_idx).border = bot_border
        ws.merge_cells(start_row=label_row, start_column=7, end_row=label_row, end_column=8)
        sign = ws.cell(row=label_row, column=7, value='(підпис)')
        sign.alignment = center_align

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f'{safe_category}_інвентаризація_БД.xlsx'
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Utilities ─────────────────────────────────────────────────────────────────

@app.route('/db/clear', methods=['POST'])
@login_required
def db_clear():
    """Очищає всі таблиці БД. Вимагає підтверджувальний токен."""
    token = request.json.get('confirm_token') if request.is_json else request.form.get('confirm_token')
    if token != 'CONFIRM_CLEAR':
        return jsonify({'error': 'Невірний токен підтвердження. Передайте confirm_token=CONFIRM_CLEAR'}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('TRUNCATE TABLE article_snapshots, operations, uploads, articles CASCADE')
        logging.warning('Database cleared: all tables truncated via /db/clear endpoint')
        return jsonify({'success': True, 'message': 'Всі дані видалено'})
    except Exception as e:
        logging.exception('Error clearing DB')
        return jsonify({'error': str(e)}), 500


@app.route('/backup')
@login_required
def backup():
    """Скачує CSV-дамп таблиці operations."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT o.id, o.article_id, a.name, o.doc_type, o.doc_code,
                           o.subdoc_type, o.subdoc_code, o.direction,
                           o.op_date, o.qty, o.col_source, o.import_id
                    FROM operations o
                    JOIN articles a USING (article_id)
                    ORDER BY o.op_date, o.article_id
                    """
                )
                rows = cur.fetchall()

        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['id', 'article_id', 'name', 'doc_type', 'doc_code',
                         'subdoc_type', 'subdoc_code', 'direction',
                         'op_date', 'qty', 'col_source', 'import_id'])
        for r in rows:
            writer.writerow([
                r['id'], r['article_id'], r['name'], r['doc_type'], r['doc_code'],
                r['subdoc_type'], r['subdoc_code'], r['direction'],
                r['op_date'], r['qty'], r['col_source'], r['import_id'],
            ])
        buf.seek(0)
        bytes_buf = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
        return send_file(bytes_buf, as_attachment=True,
                         download_name='operations_backup.csv',
                         mimetype='text/csv; charset=utf-8')
    except Exception as e:
        logging.exception('Error creating backup')
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
